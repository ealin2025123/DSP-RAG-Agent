import re
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path


@dataclass
class ParsedDocument:
    text: str
    source_type: str
    metadata: dict


class LoaderError(RuntimeError):
    pass


class TextLoader:
    def load(self, path):
        for encoding in ("utf-8-sig", "utf-8", "gb18030"):
            try:
                return ParsedDocument(Path(path).read_text(encoding=encoding), Path(path).suffix.lower()[1:], {})
            except UnicodeDecodeError:
                continue
        raise LoaderError("无法识别文本编码")


class DocxLoader:
    def load(self, path):
        try:
            with zipfile.ZipFile(str(path)) as archive:
                root = ET.fromstring(archive.read("word/document.xml"))
        except (KeyError, zipfile.BadZipFile, ET.ParseError) as exc:
            raise LoaderError("DOCX 解析失败: {}".format(exc))
        lines = []
        for paragraph in root.iter():
            if paragraph.tag.endswith("}p"):
                texts = [node.text or "" for node in paragraph.iter() if node.tag.endswith("}t")]
                text = "".join(texts).strip()
                if text:
                    lines.append(text)
        return ParsedDocument("\n\n".join(lines), "docx", {"paragraphs": len(lines)})


class PptxLoader:
    SLIDE = re.compile(r"ppt/slides/slide(\d+)\.xml$")

    def load(self, path):
        sections = []
        try:
            with zipfile.ZipFile(str(path)) as archive:
                slides = []
                for name in archive.namelist():
                    match = self.SLIDE.match(name)
                    if match:
                        slides.append((int(match.group(1)), name))
                for number, name in sorted(slides):
                    root = ET.fromstring(archive.read(name))
                    texts = [(node.text or "").strip() for node in root.iter() if node.tag.endswith("}t")]
                    texts = [text for text in texts if text]
                    if texts:
                        sections.append("## 幻灯片 {}\n\n{}".format(number, "\n\n".join(texts)))
        except (zipfile.BadZipFile, ET.ParseError) as exc:
            raise LoaderError("PPTX 解析失败: {}".format(exc))
        return ParsedDocument("\n\n".join(sections), "pptx", {"slides": len(sections)})


class XlsxLoader:
    @staticmethod
    def _escape(value):
        return str(value).strip().replace("|", "\\|").replace("\r", " ").replace("\n", " / ")

    def load(self, path):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise LoaderError("读取 XLSX 需要安装 openpyxl")
        try:
            workbook = load_workbook(str(path), read_only=True, data_only=True)
        except Exception as exc:
            raise LoaderError("XLSX 解析失败: {}".format(exc))
        sections = []
        try:
            for sheet in workbook.worksheets:
                if sheet.sheet_state != "visible":
                    continue
                rows = []
                for raw in sheet.iter_rows(values_only=True):
                    values = [self._escape(value) if value is not None else "" for value in raw]
                    while values and not values[-1]:
                        values.pop()
                    if any(values):
                        rows.append(values)
                if not rows:
                    continue
                width = max(len(row) for row in rows)
                normalized = [row + [""] * (width - len(row)) for row in rows]
                header = normalized[0]
                if len(set(cell for cell in header if cell)) != len([cell for cell in header if cell]):
                    header = ["列{}".format(i + 1) for i in range(width)]
                    normalized.insert(0, header)
                table = ["| " + " | ".join(header) + " |", "| " + " | ".join(["---"] * width) + " |"]
                table.extend("| " + " | ".join(row) + " |" for row in normalized[1:])
                sections.append("## 工作表：{}\n\n{}".format(sheet.title, "\n".join(table)))
        finally:
            workbook.close()
        return ParsedDocument("\n\n".join(sections), "xlsx", {"sheets": len(sections)})


class PdfLoader:
    @staticmethod
    def _page_numbers(page_spec, total):
        if not page_spec:
            return list(range(1, total + 1))
        selected = set()
        for part in str(page_spec).split(","):
            part = part.strip()
            if not part:
                continue
            if "-" in part:
                start, end = part.split("-", 1)
                start, end = int(start), int(end)
                if start > end:
                    raise LoaderError("PDF 页码范围无效: {}".format(part))
                selected.update(range(start, end + 1))
            else:
                selected.add(int(part))
        if not selected or min(selected) < 1 or max(selected) > total:
            raise LoaderError("PDF 页码超出范围（共 {} 页）".format(total))
        return sorted(selected)

    def load(self, path, pages=None):
        try:
            from PyPDF2 import PdfReader
        except ImportError:
            raise LoaderError("读取 PDF 需要安装 PyPDF2==3.0.1")
        try:
            reader = PdfReader(str(path))
            sections = []
            page_numbers = self._page_numbers(pages, len(reader.pages))
            for index in page_numbers:
                page = reader.pages[index - 1]
                text = (page.extract_text() or "").strip()
                if text:
                    sections.append("## 第 {} 页\n\n{}".format(index, text))
        except Exception as exc:
            raise LoaderError("PDF 解析失败: {}".format(exc))
        return ParsedDocument(
            "\n\n".join(sections),
            "pdf",
            {"total_pages": len(reader.pages), "selected_pages": page_numbers, "text_pages": len(sections)},
        )


class LoaderFactory:
    LOADERS = {
        ".txt": TextLoader,
        ".md": TextLoader,
        ".docx": DocxLoader,
        ".pptx": PptxLoader,
        ".xlsx": XlsxLoader,
        ".pdf": PdfLoader,
    }

    @classmethod
    def load(cls, path, pages=None):
        path = Path(path).resolve()
        if not path.exists() or not path.is_file():
            raise LoaderError("文件不存在: {}".format(path))
        loader_class = cls.LOADERS.get(path.suffix.lower())
        if loader_class is None:
            raise LoaderError("不支持的格式: {}".format(path.suffix))
        if pages and path.suffix.lower() != ".pdf":
            raise LoaderError("--pages 仅适用于 PDF")
        document = loader_class().load(path, pages=pages) if path.suffix.lower() == ".pdf" else loader_class().load(path)
        if not document.text.strip():
            raise LoaderError("未提取到可用文本；扫描 PDF 或图片需要视觉/OCR Loader")
        return document
